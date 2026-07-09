"""소재매핑 탭 시드 (1회성) — 로컬 엑셀의 수동 매핑을 시트에 '수동'으로 적재.

엑셀은 첫 열=광고이름, 둘째 열=소재 (기본 시트명 'Sheet1').
실행:
  PLAUD_SHEET_ID=... GOOGLE_APPLICATION_CREDENTIALS=/경로/sa.json \
  python -m collectors.seed_creative_map "/경로/일별-구매-링크 (8).xlsx" [시트명]
"""
import sys

import openpyxl
import pandas as pd

from collectors import config, sheets_io


def main(path: str, sheet: str = "Sheet1") -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    seen: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        ad, so = ws.cell(r, 1).value, ws.cell(r, 2).value
        if ad is None:
            continue
        ad = str(ad).strip()
        so = "" if so is None else str(so).strip()
        if ad and so:
            seen[ad] = so  # 중복 광고이름은 마지막 값
    rows = [{"광고이름": ad, "소재": so, "분류방식": "수동"} for ad, so in seen.items()]
    df = pd.DataFrame(rows, columns=config.CREATIVE_MAP_COLUMNS)
    n = sheets_io.write_tab(df, config.TAB_CREATIVE_MAP, config.CREATIVE_MAP_COLUMNS)
    print(f"소재매핑 시드 완료: {n}행 (수동) · distinct 소재 {df['소재'].nunique()}개")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("사용법: python -m collectors.seed_creative_map <엑셀경로> [시트명]")
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "Sheet1")
