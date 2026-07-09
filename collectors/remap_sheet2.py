"""소재매핑을 Sheet2 목록(어휘) 기준으로 재구성 (1회성, 로컬).

- Sheet1(광고이름→소재)의 각 광고를 Sheet2 어휘에 맞춰 재배치:
  normalize(광고이름)가 Sheet2 목록에 있으면 그것, 아니면 Sheet1 값.
- meta_소재일별의 광고이름 중 위 매핑에 없는 건 '제외'.

실행:
  PLAUD_SHEET_ID=... GOOGLE_APPLICATION_CREDENTIALS=/경로/sa.json \
  python -m collectors.remap_sheet2 "/경로/일별-구매-링크 (8).xlsx"
"""
import sys

import openpyxl
import pandas as pd

from collectors import config, creative_mapping, sheets_io


def main(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws1, ws2 = wb["Sheet1"], wb["Sheet2"]

    sheet1 = {}
    for r in range(2, ws1.max_row + 1):
        ad, so = ws1.cell(r, 1).value, ws1.cell(r, 2).value
        if ad is None or so is None:
            continue
        sheet1[str(ad).strip()] = str(so).strip()

    s2 = {str(ws2.cell(r, 7).value).strip() for r in range(1, ws2.max_row + 1)
          if ws2.cell(r, 7).value not in (None, "")}

    manual = {}
    for ad, so1 in sheet1.items():
        cand = creative_mapping.normalize(ad)
        manual[ad] = cand if cand in s2 else so1

    meta = sheets_io.read_tab(config.TAB_META_CREATIVE_DAILY, config.META_CREATIVE_DAILY_COLUMNS)
    ad_names = set(meta["ad_name"].astype(str).str.strip()) if not meta.empty else set()

    rows = [{"광고이름": a, "소재": s, "분류방식": "수동"} for a, s in manual.items()]
    for a in sorted(ad_names - set(manual)):
        if a and a != "nan":
            rows.append({"광고이름": a, "소재": creative_mapping.normalize(a), "분류방식": "제외"})

    df = pd.DataFrame(rows, columns=config.CREATIVE_MAP_COLUMNS)
    n = sheets_io.write_tab(df, config.TAB_CREATIVE_MAP, config.CREATIVE_MAP_COLUMNS)
    man = sum(1 for r in rows if r["분류방식"] == "수동")
    print(f"재구성 완료 — 수동 {man}건 / distinct 소재 {len(set(manual.values()))} · 제외 {n - man} · 총 {n}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("사용법: python -m collectors.remap_sheet2 <엑셀경로>")
    main(sys.argv[1])
