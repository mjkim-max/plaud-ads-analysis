"""소재매핑 전면 재구성 (1회성, 로컬) — 전 기간 데이터를 사용자 파일의 소재명으로.

우선순위: Sheet1(정리본) > Raw Data Report 소재명(원본, 초기 포함) > normalize(폴백).
정크(테스트·장바구니·결제시작·빈값)만 '제외', 나머지는 전부 '수동'으로 표시.

실행:
  PLAUD_SHEET_ID=... GOOGLE_APPLICATION_CREDENTIALS=/경로/sa.json \
  python -m collectors.build_full_map "/경로/일별-구매-링크 (8).xlsx"
"""
import sys

import openpyxl
import pandas as pd

from collectors import config, creative_mapping, sheets_io

JUNK = ("test", "장바구니", "결제시작")


def _is_junk(so: str) -> bool:
    s = str(so).strip()
    if not s or s == "nan":
        return True
    low = s.lower()
    return any(k in low for k in JUNK)


def main(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)

    ws1 = wb["Sheet1"]
    sheet1 = {}
    for r in range(2, ws1.max_row + 1):
        ad, so = ws1.cell(r, 1).value, ws1.cell(r, 2).value
        if ad is not None and so is not None and str(so).strip():
            sheet1[str(ad).strip()] = str(so).strip()

    wsr = wb["Raw Data Report"]
    rdr = {}
    for r in range(2, wsr.max_row + 1):
        ad, so = wsr.cell(r, 2).value, wsr.cell(r, 7).value
        if ad is not None and so is not None and str(so).strip():
            rdr[str(ad).strip()] = str(so).strip()

    meta = sheets_io.read_tab(config.TAB_META_CREATIVE_DAILY, config.META_CREATIVE_DAILY_COLUMNS)
    ad_names = sorted(set(meta["ad_name"].astype(str).str.strip())) if not meta.empty else []

    rows = []
    for ad in ad_names:
        if not ad or ad == "nan":
            continue
        so = sheet1.get(ad) or rdr.get(ad) or creative_mapping.normalize(ad)
        how = "제외" if _is_junk(so) else "수동"
        rows.append({"광고이름": ad, "소재": so, "분류방식": how})

    df = pd.DataFrame(rows, columns=config.CREATIVE_MAP_COLUMNS)
    n = sheets_io.write_tab(df, config.TAB_CREATIVE_MAP, config.CREATIVE_MAP_COLUMNS)
    shown = df[df["분류방식"] == "수동"]
    print(f"재구성 — 총 {n}행 · 수동 {len(shown)} · 제외 {n - len(shown)} · "
          f"표시 소재 {shown['소재'].nunique()}개")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("사용법: python -m collectors.build_full_map <엑셀경로>")
    main(sys.argv[1])
